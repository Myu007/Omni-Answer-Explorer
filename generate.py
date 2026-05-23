#!/usr/bin/env python3
"""
Omni-Answer Explorer — Build Script
====================================
Scans source/ for all .xlsx/.csv files, extracts questions tagged with
semester/subject info, and injects everything into template.html.

Designed for GitHub Actions: push Excel files → auto rebuild → auto deploy.

Usage:
    python generate.py          # requires GROQ_API_KEY env var

Requirements:
    pip install openpyxl
"""

import json, os, sys, glob

# ─── Config ───────────────────────────────────────────────────────────────────

GROQ_API_KEY = os.environ.get("GROQ_API_KEY", "")

SOURCE_DIR = "source"
TEMPLATE_PATH = "template.html"
OUTPUT_PATH = "index.html"

MODELS = {
    "ai1": "llama-3.3-70b-versatile",
    "ai2": "groq/compound-mini",
    "ai3": "llama-3.1-8b-instant",
    "ai4": "meta-llama/llama-4-scout-17b-16e-instruct",
}

# ─── Helpers ──────────────────────────────────────────────────────────────────

def log(msg):
    print(f"[generate] {msg}")


def read_excel_questions(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet_name = "Bank" if "Bank" in wb.sheetnames else wb.sheetnames[0]
    sheet = wb[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    wb.close()

    header_idx = None
    q_col = 0
    for i, row in enumerate(rows):
        if row is None:
            continue
        for j, cell in enumerate(row):
            if cell is not None and isinstance(cell, str) and cell.strip().lower() == "question":
                header_idx = i
                q_col = j
                break
        if header_idx is not None:
            break

    start = header_idx + 1 if header_idx is not None else 0
    questions = []
    for row in rows[start:]:
        if row is None or q_col >= len(row):
            continue
        val = row[q_col]
        if val is not None:
            text = str(val).strip()
            if text:
                questions.append(text)
    return questions


def read_csv_questions(path):
    import csv
    with open(path, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        rows = list(reader)

    header_idx = None
    q_col = 0
    for i, row in enumerate(rows):
        if not row:
            continue
        for j, cell in enumerate(row):
            if cell is not None and cell.strip().lower() == "question":
                header_idx = i
                q_col = j
                break
        if header_idx is not None:
            break

    start = header_idx + 1 if header_idx is not None else 0
    questions = []
    for row in rows[start:]:
        if not row or q_col >= len(row):
            continue
        val = row[q_col]
        if val is not None:
            text = str(val).strip()
            if text:
                questions.append(text)
    return questions


def read_all_files(source_dir):
    """Scan source/ recursively, read all Excel/CSV files, return tagged questions."""
    all_questions = []
    subjects = []  # list of { id, semester, subject, count }

    pattern = os.path.join(source_dir, "**", "*.*")
    for fpath in sorted(glob.glob(pattern, recursive=True)):
        ext = os.path.splitext(fpath)[1].lower()
        if ext not in (".xlsx", ".xls", ".csv"):
            continue

        # Derive semester + subject from folder structure
        rel = os.path.relpath(fpath, source_dir)
        parts = rel.replace("\\", "/").split("/")
        semester = parts[0] if len(parts) > 1 else "General"
        subject = os.path.splitext(parts[-1])[0]

        log(f"  {rel}")

        try:
            if ext == ".csv":
                qs = read_csv_questions(fpath)
            else:
                qs = read_excel_questions(fpath)
        except Exception as e:
            log(f"    ✗ Error: {e}")
            continue

        for q in qs:
            all_questions.append({
                "text": q,
                "semester": semester,
                "subject": subject,
            })

        subjects.append({
            "id": f"{semester}/{subject}",
            "semester": semester,
            "subject": subject,
            "count": len(qs),
        })

    return all_questions, subjects


def build_config_block():
    config = {
        "models": MODELS,
        "maxTokens": 300,
        "temperature": 0.7,
        "maxCustomQuestionsPerSession": 3,
        "cooldownSeconds": 10,
        "groqApiKey": GROQ_API_KEY if GROQ_API_KEY else "",
    }
    return json.dumps(config, indent=2)


def inject_into_template(questions, subjects, template_path, output_path):
    with open(template_path, "r", encoding="utf-8") as f:
        html = f.read()

    questions_json = json.dumps(questions, ensure_ascii=False)
    subjects_json = json.dumps(subjects, ensure_ascii=False)
    config_json = build_config_block()

    html = html.replace("/* INJECT_QUESTIONS */", questions_json)
    html = html.replace("/* INJECT_SUBJECTS */", subjects_json)
    html = html.replace("/* INJECT_CACHE */", "{}")
    html = html.replace("/* INJECT_CONFIG */", config_json)

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)

    size_kb = os.path.getsize(output_path) / 1024
    log(f"Written {output_path} ({size_kb:.1f} KB)")


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    if not os.path.exists(TEMPLATE_PATH):
        log(f"ERROR: Template not found at {TEMPLATE_PATH}")
        sys.exit(1)

    if not os.path.exists(SOURCE_DIR):
        log(f"Creating {SOURCE_DIR}/ — place your .xlsx/.csv files there.")
        os.makedirs(SOURCE_DIR)

    log(f"Scanning {SOURCE_DIR}/…")
    questions, subjects = read_all_files(SOURCE_DIR)

    total_q = len(questions)
    total_subjects = len(subjects)
    log(f"Found {total_q} questions across {total_subjects} subjects")

    if total_q == 0:
        log("WARNING: No questions found. The site will be empty.")
        log(f"Place .xlsx or .csv files with a 'Question' column in {SOURCE_DIR}/")

    log(f"Injecting into {TEMPLATE_PATH}…")
    inject_into_template(questions, subjects, TEMPLATE_PATH, OUTPUT_PATH)

    if not GROQ_API_KEY:
        log("WARNING: GROQ_API_KEY is empty — custom questions and live AI answers won't work.")
        log("Set GROQ_API_KEY as an environment variable or edit this script.")

    log(f"Done. {OUTPUT_PATH} ready to deploy.")

    if os.name == "nt":  # only pause on Windows when run interactively
        try:
            input("\nPress Enter to exit...")
        except EOFError:
            pass


if __name__ == "__main__":
    main()
